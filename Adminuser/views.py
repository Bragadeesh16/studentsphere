from django.shortcuts import render, redirect
from django.http import HttpResponse, Http404
import os
from .forms import *
from django.contrib import messages
from django.contrib.auth.models import Group
from django.forms import *
from myapp.decorators import (
    authenticate_staff,
    authenticate_users,
)
from django.conf import settings
from django.http import HttpResponse, Http404
from account.models import CustomUser, UserProfile
from .models import *
import openpyxl




def student_details(request):
    departments = [
        "Computer science",
        "Civil",
        "Mechanical",
        "ECE",
        "EEE",
        "IT",
        "Cyber security",
        "AIDS",
    ]
    years = ["First year", "Second year", "Third year", "Final year"]
    
    department_year_data = {}

    for department in departments:
        department_year_data[department] = {}
        for year in years:
            students = UserProfile.objects.filter(
                department=department, year=year
            )
            department_year_data[department][year] = students

    
    return render(request, "student_details.html", {"department_year_data": department_year_data})



@authenticate_users
def notes_folder(request, pk):
    group_name = ClassGroup.objects.get(pk=pk)
    permission = False
    if request.user.groups.filter(name="staff").exists():
        permission = True
    if request.method == "POST":
        form = uploding_folder(request.POST)
        if form.is_valid():
            instance = form.save(commit=False)
            instance.folder_from = ClassGroup.objects.get(name=group_name)
            instance.save()
        else:
            print("Your form is not valid")

    if group_name:
        folders_list = ClassGroup.objects.get(name=group_name)
        print(folders_list)
        linked_folders = folders.objects.filter(folder_from=folders_list)
    else:
        linked_folders = []

    form = uploding_folder()

    return render(
        request,
        "notes_folder.html",
        {
            "form": form,
            "group_name": group_name,
            "folder_name": linked_folders,
            "pk": pk,
            "permission": permission,
        },
    )


@authenticate_users
def notes(request, group_id, folder_id):
    form = uploding_documents()
    group = ClassGroup.objects.get(pk=group_id)
    folder = folders.objects.get(pk=folder_id)
    permission = False

    if request.user.groups.filter(name="staff").exists():
        permission = True

    if request.method == "POST":
        form = uploding_documents(request.POST, request.FILES)
        if form.is_valid():
            instance = form.save(commit=False)
            instance.file_from = folder
            instance.save()
            messages.success(
                request,
                "Your file is uploaded successfully",
            )
            return redirect("notes", group_id, folder_id)
    try:
        files_list = file_uplode.objects.filter(file_from=folder)
    except folders.DoesNotExist:
        raise Http404("Folder does not exist")
    except file_uplode.DoesNotExist:
        files_list = []

    return render(
        request,
        "notes.html",
        {
            "notes": form,
            "file_list": files_list,
            "permission": permission,
            "group_id": group_id,
            "folder_id": folder_id,
        },
    )


def delete_folder(request):
    if request.method == "POST":
        folder_id, pk = request.POST.get("delete_folder_name").split()
        folder_name = folders.objects.get(id=folder_id)
        folder_name.delete()
        return redirect("folder-notes", pk)


def delete_file(request, group_id, folder_id):
    if request.method == "POST":
        file_id = request.POST.get("file_id")
        print(file_id)
        file = file_uplode.objects.get(id=file_id)
        file.delete()
        return redirect("notes", group_id, folder_id)


def folder_rename(request):
    if request.method == "POST":
        file_id, pk = request.POST.get("folder_id").split()
        rename = request.POST.get("rename")
        folder = folders.objects.get(id=file_id)
        folder.name = rename
        folder.save()
        return redirect("folder-notes", pk)


def download(request, file_id):
    try:
        file = file_uplode.objects.get(pk=file_id)
        path = file.notes
        file_path = os.path.join(settings.MEDIA_ROOT, str(path))

        print("File Path:", file_path) 

        if os.path.exists(file_path):
            with open(file_path, "rb") as fh:
                response = HttpResponse(
                    fh.read(),
                    content_type="application/vnd.ms-excel",
                )
                response["Content-Disposition"] = (
                    "inline; filename=" + os.path.basename(file_path)
                )
                return response
        else:
            print("File does not exist") 
            raise Http404("File not found")
    except file_uplode.DoesNotExist:
        raise Http404("File does not exist")
    except Exception as e:
        print(f"An error occurred: {str(e)}")
        raise Http404("An error occurred while processing the request")



def export_student_data(request):
    departments = [
        "Computer science",
        "Civil",
        "Mechanical",
        "ECE",
        "EEE",
        "IT",
        "Cyber security",
        "AIDS",
    ]
    years = ["First year", "Second year", "Third year", "Final year"]
    workbook = openpyxl.Workbook()
    for department in departments:
        for year in years:
            students = UserProfile.objects.filter(department=department, year=year)
            if students.exists():
                sheet = workbook.create_sheet(title=f"{department} - {year}")
                headers = [
                    "Name", "Phone Number", "Gender", "Department", "Year", 
                    "Address", "Aadhar Number", "Father Name", "Mother Name", 
                    "Father Occupation", "Mother Occupation", 
                    "Father Phone Number", "Mother Phone Number", 
                    "Annual Income", "Religion", "Caste", 
                    "Community", "Mother Language"
                ]
                sheet.append(headers)

                for student in students:
                    row = [
                        student.name, student.phone_number, student.gender, 
                        student.department, student.year, 
                        student.address, student.aadhar_number, 
                        student.father_name, student.mother_name, 
                        student.father_occupation, student.mother_occupation, 
                        student.father_phone_number, student.mother_phone_number, 
                        student.annual_income, student.religion, 
                        student.caste, student.community, 
                        student.mother_language
                    ]
                    sheet.append(row)

    if 'Sheet' in workbook.sheetnames:
        del workbook['Sheet']

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=student_data.xlsx'
    workbook.save(response)
    return response